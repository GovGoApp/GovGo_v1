# =======================================================================
# PROCESSAMENTO DE ITENS DE CONTRATAÇÕES DO PNCP - PIPELINE GOVGO
# =======================================================================
# Este script processa os itens das contratações baixadas do PNCP,
# extraindo informações detalhadas de cada item contratado através
# da API do PNCP.
# 
# Funcionalidades:
# - Leitura automática dos arquivos de contratações do dia
# - Extração paralela de itens via API do PNCP
# - Processamento robusto com controle de erros
# - Organização estruturada dos dados de itens
# - Logs detalhados e barra de progresso
# 
# Resultado: Dados de itens estruturados e prontos para importação no BD.
# =======================================================================

import requests
import pandas as pd
import os
import datetime
import concurrent.futures
import re
import json
import threading
import sys
import glob
from rich.progress import Progress, BarColumn, TextColumn, TimeElapsedColumn, TimeRemainingColumn, SpinnerColumn
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from dotenv import load_dotenv

# Configuração do console Rich
console = Console()

def log_message(message, log_type="info"):
    """Log padronizado para o pipeline"""
    if log_type == "info":
        console.print(f"[white]ℹ️  {message}[/white]")
    elif log_type == "success":
        console.print(f"[green]✅ {message}[/green]")
    elif log_type == "warning":
        console.print(f"[yellow]⚠️  {message}[/yellow]")
    elif log_type == "error":
        console.print(f"[red]❌ {message}[/red]")
    elif log_type == "debug":
        console.print(f"[cyan]🔧 {message}[/cyan]")


# Obter a data atual para nomear os arquivos
today = datetime.date.today()
data_str = today.strftime("%Y%m%d")
#data_str = '20250526'  # Para testes, use uma data fixa

# Carregar configurações de caminhos
script_dir = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(script_dir, "files.env"))

# Caminhos e nomes de arquivos baseados na data atual
# Usar caminhos do arquivo de configuração
IN_PATH = os.getenv("CONTRATACOES_NEW")
OUT_PATH = os.getenv("ITENS_NEW")

OUT_FILE = f"ITENS_CONTRATAÇÕES_PNCP_{data_str}.xlsx"

# Variáveis globais para escrita imediata
workbook_lock = threading.Lock()
# Cria um novo workbook e remove a folha padrão
wb = Workbook()
wb.remove(wb.active)

# Lista de campos dos itens conforme o exemplo do arquivo itens.json (fixa)
ITEM_FIELDS = [
    "numeroItem",
    "descricao",
    "materialOuServico",
    "materialOuServicoNome",
    "valorUnitarioEstimado",
    "valorTotal",
    "quantidade",
    "unidadeMedida",
    "orcamentoSigiloso",
    "itemCategoriaId",
    "itemCategoriaNome",
    "patrimonio",
    "codigoRegistroImobiliario",
    "criterioJulgamentoId",
    "criterioJulgamentoNome",
    "situacaoCompraItem",
    "situacaoCompraItemNome",
    "tipoBeneficio",
    "tipoBeneficioNome",
    "incentivoProdutivoBasico",
    "dataInclusao",
    "dataAtualizacao",
    "temResultado",
    "imagem",
    "aplicabilidadeMargemPreferenciaNormal",
    "aplicabilidadeMargemPreferenciaAdicional",
    "percentualMargemPreferenciaNormal",
    "percentualMargemPreferenciaAdicional",
    "ncmNbsCodigo",
    "ncmNbsDescricao",
    "catalogo",
    "categoriaItemCatalogo",
    "catalogoCodigoItem",
    "informacaoComplementar"
]

# Arquivo de controle para rastrear arquivos processados
PROCESSED_FILE = os.path.join(script_dir, "processed_contratacoes_itens.log")

def remove_illegal_chars(value):
    if isinstance(value, str):
        return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', value)
    return value

def clean_dataframe(df):
    return df.apply(lambda col: col.map(remove_illegal_chars))

# Processa uma linha usando apenas o campo "numeroControlePNCP"
def process_row(row):
    numeroControle = str(row.get("numeroControlePNCP", "")).strip()
    if not numeroControle:
        return []
    try:
        # Tentar reconhecer o formato pelo padrão esperado
        if not re.match(r'^\d+-\d+-\d+/\d+$', numeroControle):
            # Formato inválido, pular silenciosamente
            return []
            
        parts = numeroControle.split("-")
        if len(parts) != 3:
            return []
        cnpj = parts[0]
        # Ignora o prefixo (parts[1])
        seq_and_year = parts[2].split("/")
        if len(seq_and_year) != 2:
            return []
        seq = seq_and_year[0]
        anoCompra = seq_and_year[1]
        # Remove zeros à esquerda do sequencial
        sequencialCompra = str(int(seq))
        # Monta a URL:
        url = f"https://pncp.gov.br/api/pncp//v1/orgaos/{cnpj}/compras/{anoCompra}/{sequencialCompra}/itens"
        # Definindo timeout para evitar travamentos
        response = requests.get(url, timeout=20)
        if response.status_code != 200:
            return []
        itens = response.json()  # espera-se que seja uma lista
        resultados = []
        for item in itens:
            registro = {"numeroControlePNCP": numeroControle}
            for campo in ITEM_FIELDS:
                valor = item.get(campo)
                # Converter dicionários e listas para string JSON
                if isinstance(valor, (dict, list)):
                    registro[campo] = json.dumps(valor, ensure_ascii=False)
                else:
                    registro[campo] = valor
            resultados.append(registro)
        return resultados
    except requests.exceptions.Timeout:
        log_message(f"Timeout ao processar {numeroControle}", "warning")
        return []
    except Exception as e:
        log_message(f"Erro ao processar {numeroControle}: {str(e)}", "error")
        return [{"numeroControlePNCP": numeroControle, "erro": str(e)}]
    
# Processa uma aba (sheet) do IN_FILE
def process_sheet(sheet_name, df, output_file, progress):
    # Remove linhas totalmente vazias
    df = df.dropna(how="all")
    log_message(f"Aba '{sheet_name}' lida com {len(df)} linhas", "info")
    
    total_rows = len(df)
    sheet_task_id = progress.add_task(f"[bold blue]📄 Aba {sheet_name}", total=total_rows)
    
    # Dividir o processamento em blocos de 1000 linhas
    BATCH_SIZE = 1000
    resultados_totais = 0
    
    for start_idx in range(0, total_rows, BATCH_SIZE):
        end_idx = min(start_idx + BATCH_SIZE, total_rows)
        batch_df = df.iloc[start_idx:end_idx]
        batch_size = len(batch_df)
        
        log_message(f"Processando batch {start_idx//BATCH_SIZE + 1} ({start_idx+1}-{end_idx})", "debug")
        
        # Adicionar barra para o batch atual
        batch_id = progress.add_task(
            f"[bold yellow]  📦 Batch {start_idx//BATCH_SIZE + 1} ({start_idx+1}-{end_idx})", 
            total=batch_size
        )
        
        resultados = []
        
        # Processar o lote atual
        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            future_to_index = {executor.submit(process_row, row): idx for idx, row in batch_df.iterrows()}
            for future in concurrent.futures.as_completed(future_to_index):
                try:
                    linhas = future.result()
                    for linha in linhas:
                        if 'erro' in linha:
                            log_message(f"Erro linha '{linha['numeroControlePNCP']}': {linha['erro']}", "error")
                        else:
                            resultados.append(linha)
                except Exception as e:
                    log_message(f"Erro inesperado na execução da thread: {str(e)}", "error")
                
                # Atualizar ambas as barras de progresso
                progress.update(batch_id, advance=1)
                progress.update(sheet_task_id, advance=1)
        
        # Criar DataFrame para o lote atual
        colunas_saida = ["numeroControlePNCP"] + ITEM_FIELDS
        if resultados:
            batch_result_df = pd.DataFrame(resultados)
            batch_result_df = clean_dataframe(batch_result_df)
            batch_result_df = batch_result_df.reindex(columns=colunas_saida)
            
            # Salvar o lote atual e liberar memória
            append_to_sheet(sheet_name, batch_result_df, output_file, progress)
            resultados_totais += len(batch_result_df)
            
            # Remover a barra do batch concluído
            progress.update(batch_id, completed=batch_size, 
                description=f"[bold green]  📦 Batch {start_idx//BATCH_SIZE + 1:02d} ✓ {len(batch_result_df)} itens")
            
            log_message(f"Batch {start_idx//BATCH_SIZE + 1} concluído: {len(batch_result_df)} itens", "success")
            
            # Liberar memória
            del batch_result_df
            del resultados
        else:
            progress.remove_task(batch_id)
            log_message(f"Batch {start_idx//BATCH_SIZE + 1}: Sem resultados", "warning")
    
    progress.remove_task(sheet_task_id)
    log_message(f"Concluída a aba {sheet_name} com {resultados_totais} itens", "success")
    
    # Retornar uma estrutura vazia apenas para manter compatibilidade
    return pd.DataFrame(columns=colunas_saida)

# Função para acrescentar dados em uma aba sem recarregar todo o arquivo
def append_to_sheet(sheet_name, df, output_file, progress):
    """Adiciona dados a um arquivo Excel, criando-o se não existir"""
    try:
        # Cria diretório de saída se não existir
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        
        # Verifica se o arquivo existe
        if not os.path.exists(output_file):
            # Se não existe, cria um novo arquivo
            temp_wb = Workbook()
            # Sempre remover a primeira aba em branco
            if "Sheet" in temp_wb.sheetnames:
                temp_wb.remove(temp_wb["Sheet"])
            ws = temp_wb.create_sheet(title=sheet_name)
            headers_added = False
        else:
            # Se existe, carrega o arquivo
            try:
                temp_wb = openpyxl.load_workbook(output_file)
                if sheet_name in temp_wb.sheetnames:
                    ws = temp_wb[sheet_name]
                    headers_added = True
                else:
                    ws = temp_wb.create_sheet(title=sheet_name)
                    headers_added = False
            except Exception as e:
                progress.console.log(f"[red]Erro ao carregar arquivo para append: {str(e)}")
                # Criar um novo arquivo se falhar ao carregar
                temp_wb = Workbook()
                if "Sheet" in temp_wb.sheetnames:
                    temp_wb.remove(temp_wb["Sheet"])
                ws = temp_wb.create_sheet(title=sheet_name)
                headers_added = False
        
        # Adiciona dados
        for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True)):
            if r_idx == 0 and headers_added:
                continue  # Pular cabeçalhos se já existirem
            ws.append(r)
            
        # Salva o arquivo
        temp_wb.save(output_file)
        progress.console.log(f"[green]Dados salvos em: {output_file}, aba: {sheet_name}")
            
    except Exception as e:
        progress.console.log(f"[red]Erro ao salvar arquivo {output_file}: {str(e)}")
        # Tenta salvar em um arquivo alternativo
        backup_file = output_file.replace('.xlsx', f'_backup_{datetime.datetime.now().strftime("%H%M%S")}.xlsx')
        progress.console.log(f"[yellow]Tentando salvar em arquivo alternativo: {backup_file}")
        try:
            temp_wb.save(backup_file)
        except Exception as backup_err:
            progress.console.log(f"[bold red]Falha também no backup: {str(backup_err)}")

def get_processed_files():
    """Retorna a lista de arquivos já processados"""
    if os.path.exists(PROCESSED_FILE):
        with open(PROCESSED_FILE, 'r') as f:
            return set(line.strip() for line in f if line.strip())
    return set()

def mark_as_processed(filename):
    """Marca um arquivo como processado"""
    with open(PROCESSED_FILE, 'a') as f:
        f.write(f"{filename}\n")

def main():
    console.print(Panel("[bold blue] [2/7] PROCESSAMENTO DE ITENS PNCP[/bold blue]"))
    
    # Listar arquivos de contratações para processar
    input_files = glob.glob(os.path.join(IN_PATH, "CONTRATAÇÕES_PNCP_*.xlsx"))
    processed_files = get_processed_files()
    files_to_process = [f for f in input_files if os.path.basename(f) not in processed_files]
    
    if not files_to_process:
        log_message("Todos os arquivos já foram processados", "success")
        return
    
    log_message(f"Processando {len(files_to_process)} arquivo(s)")
    
    with Progress(
        SpinnerColumn(),
        TextColumn("[bold blue]{task.description}"),
        BarColumn(),
        "[progress.percentage]{task.percentage:>3.0f}%",
        TimeElapsedColumn(),
        console=console
    ) as progress:
        main_task = progress.add_task("Processamento de itens", total=len(files_to_process))
        
        for input_file in files_to_process:
            filename = os.path.basename(input_file)
            
            # Extrair a data do nome do arquivo
            date_match = re.search(r'CONTRATAÇÕES_PNCP_(\d{8})\.xlsx', filename)
            if not date_match:
                log_message(f"Formato inválido: {filename}", "warning")
                progress.advance(main_task)
                continue
                    
            data_str = date_match.group(1)
            OUT_FILE = f"ITENS_CONTRATAÇÕES_PNCP_{data_str}.xlsx"
            
            # Resetar workbook para cada arquivo
            global wb
            wb = Workbook()
            wb.remove(wb.active)
            
            try:
                if not os.path.exists(input_file):
                    log_message(f"Arquivo não encontrado: {filename}", "error")
                    progress.advance(main_task)
                    continue
                
                # Lê o arquivo Excel com todas as abas
                xls = pd.ExcelFile(input_file)
                sheets = xls.sheet_names
                
                # Processar cada aba
                for sheet in sheets:
                    try:
                        df_sheet = pd.read_excel(xls, sheet_name=sheet)
                        output_file = os.path.join(OUT_PATH, OUT_FILE)
                        process_sheet(sheet, df_sheet, output_file, progress)
                        del df_sheet
                    except Exception as e:
                        log_message(f"Erro na aba {sheet}: {str(e)}", "error")
                
                mark_as_processed(filename)
                log_message(f"Arquivo processado: {filename}", "success")
                
            except Exception as e:
                log_message(f"Erro no arquivo {filename}: {str(e)}", "error")
            
            progress.advance(main_task)
    
    log_message("Processamento de itens concluído", "success")

if __name__ == "__main__":
    main()