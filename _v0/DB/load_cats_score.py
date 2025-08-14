# update_contratacao_categoria.py

import sqlite3
from openpyxl import load_workbook
from tqdm import tqdm




# ========== CONFIGURAÇÃO ==========
BASE_PATH = "C:\\Users\\Haroldo Duraes\\Desktop\\GOvGO\\v0\\#DATA\\PNCP\\"

DB_PATH = BASE_PATH + "DB\\"
DB_FILE = DB_PATH + 'pncp_v2.db'                            # caminho para seu banco SQLite

EXCEL_PATH = BASE_PATH + "CLASSY\\OUTPUT\\NEW_ORDER\\"      # caminho para o arquivo Excel
EXCEL_FILE = EXCEL_PATH + "UNIFIED_OUTPUT.xlsx"             # caminho para o arquivo Excel
SHEET_NAME = 'OUTPUT'                                      # aba a ser lida


# ======= PASSO 1: Abre conexão e adiciona colunas =======
conn = sqlite3.connect(DB_FILE)
cur  = conn.cursor()

for col, col_type in [('CODCAT', 'TEXT'), ('SCORE', 'REAL')]:
    try:
        cur.execute(f"ALTER TABLE contratacao ADD COLUMN {col} {col_type};")
    except sqlite3.OperationalError:
        # provavelmente já existe
        pass

conn.commit()

# ======= PASSO 2: Abre o Excel em modo read-only =======
wb = load_workbook(filename= EXCEL_FILE, read_only=True, data_only=True)
ws = wb[SHEET_NAME] if SHEET_NAME else wb[wb.sheetnames[0]]

# ======= PASSO 3: Identifica indices das colunas =======
header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
idx_id     = header.index('id_pncp')
idx_top1   = header.index('TOP_1')
idx_score1 = header.index('SCORE_1')

# ======= PASSO 4: Prepara SQL de update =======
sql_update = """
UPDATE contratacao
   SET CODCAT = ?,
       SCORE  = ?
 WHERE numeroControlePNCP = ?
;"""

# ======= PASSO 5: Itera linhas com progresso =======
batch_size = 5000
batch = []

for row in tqdm(ws.iter_rows(min_row=2), desc="Atualizando contratacao", unit="lin"):
    id_pncp = row[idx_id].value
    top1    = row[idx_top1].value
    score1  = row[idx_score1].value

    # Garante que não seja None
    if id_pncp is None:
        continue

    batch.append((str(top1), float(score1) if score1 is not None else None, str(id_pncp)))

    # Quando atinge batch_size, executa e limpa
    if len(batch) >= batch_size:
        cur.executemany(sql_update, batch)
        conn.commit()
        batch.clear()

# Insere o que sobrou
if batch:
    cur.executemany(sql_update, batch)
    conn.commit()

# ======= FINALIZA =======
conn.close()
print("✔ Atualização concluída.")
