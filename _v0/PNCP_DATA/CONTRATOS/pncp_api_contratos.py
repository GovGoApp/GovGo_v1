import requests
import pandas as pd
from tqdm import tqdm
import time
from datetime import datetime
import os
import openpyxl
from openpyxl import load_workbook

def fetch_pncp_data(data_inicial, data_final, pagina=1, tamanho_pagina=100):
    """Função para buscar dados da API do PNCP"""
    url = "https://pncp.gov.br/api/consulta/v1/contratos"
    params = {
        "dataInicial": data_inicial,
        "dataFinal": data_final,
        "pagina": pagina,
        "tamanhoPagina": tamanho_pagina
    }
    
    max_retries = 3
    retry_delay = 5  # segundos
    
    for attempt in range(max_retries):
        try:
            response = requests.get(url, params=params, timeout=30)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            print(f"Erro na requisição (tentativa {attempt+1}/{max_retries}): {e}")
            if attempt < max_retries - 1:
                print(f"Tentando novamente em {retry_delay} segundos...")
                time.sleep(retry_delay)
            else:
                print(f"Falha após {max_retries} tentativas")
                return None

def get_all_contracts_for_year(year):
    """Função para obter todos os contratos de um ano específico"""
    data_inicial = f"{year}0101"
    data_final = f"{year}1231"
    
    # Primeira requisição para obter informações de paginação
    print(f"Buscando dados para o ano {year}...")
    first_page = fetch_pncp_data(data_inicial, data_final)
    
    if not first_page or first_page.get("empty", True):
        print(f"Nenhum dado encontrado para o ano {year}")
        return []
    
    total_paginas = first_page.get("totalPaginas", 0)
    total_registros = first_page.get("totalRegistros", 0)
    print(f"Total de {total_registros} registros em {total_paginas} páginas")
    
    all_data = first_page.get("data", [])
    
    # Buscar as páginas restantes
    if total_paginas > 1:
        for pagina in tqdm(range(2, total_paginas + 1), desc=f"Páginas do ano {year}"):
            response_data = fetch_pncp_data(data_inicial, data_final, pagina)
            if response_data and not response_data.get("empty", True):
                all_data.extend(response_data.get("data", []))
            time.sleep(0.5)  # Pequeno delay para evitar sobrecarga na API
    
    return all_data

def normalize_data(contratos):
    """Função para normalizar os dados aninhados para formato plano"""
    normalized_data = []
    
    for contrato in contratos:
        flat_contrato = {}
        
        # Copia campos simples
        for key, value in contrato.items():
            if isinstance(value, (str, int, float, bool)) or value is None:
                flat_contrato[key] = value
        
        # Processa campos aninhados
        nested_fields = ["tipoContrato", "orgaoEntidade", "categoriaProcesso", 
                         "unidadeOrgao", "unidadeSubRogada", "orgaoSubRogado"]
        
        for field in nested_fields:
            if field in contrato and contrato[field]:
                for subkey, subvalue in contrato[field].items():
                    flat_contrato[f"{field}_{subkey}"] = subvalue
        
        normalized_data.append(flat_contrato)
    
    return normalized_data


def save_to_excel_in_batches(df, path, filename, batch_size=1000):
    """
    Salva o DataFrame em um único arquivo Excel, atualizando a mesma aba ('Contratos')
    a cada batch de 1000 linhas (10 páginas).
    """
    # Garantir que o diretório existe
    os.makedirs(path, exist_ok=True)
    
    # Caminho completo do arquivo
    file_path = os.path.join(path, filename)
    
    total_rows = len(df)
    num_batches = (total_rows + batch_size - 1) // batch_size
    
    for i in tqdm(range(num_batches), desc="Salvando dados em batches"):
        start_idx = i * batch_size
        end_idx = min(start_idx + batch_size, total_rows)
        batch_df = df.iloc[start_idx:end_idx]
        
        if i == 0:
            # Cria o arquivo Excel com a aba "Contratos", incluindo o cabeçalho
            batch_df.to_excel(file_path, index=False, sheet_name='Contratos')
            print(f"Arquivo Excel criado: {file_path} com linhas {start_idx+1} a {end_idx}")
        else:
            # Abre o workbook existente e obtém a aba "Contratos"
            book = load_workbook(file_path)
            if "Contratos" in book.sheetnames:
                sheet = book["Contratos"]
                start_row = sheet.max_row + 1  # Inicia logo após a última linha preenchida
            else:
                # Se a aba não existir, cria-a
                book.create_sheet("Contratos")
                start_row = 1
            
            # Anexa os novos dados sem escrever o cabeçalho (header=False)
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                writer.book = book
                batch_df.to_excel(writer, index=False, sheet_name='Contratos', startrow=start_row, header=False)
                writer.save()
            
            print(f"Batch {i+1}/{num_batches} salvo. Linhas {start_idx+1} a {end_idx}")

def main():
    # Determinar o ano atual
    current_year = datetime.now().year
    
    # Anos a serem consultados (de 2021 até o ano atual)
    years = list(range(2021, current_year + 1))
    
    all_contracts = []
    
    # Buscar dados para cada ano
    for year in years:
        contracts = get_all_contracts_for_year(year)
        all_contracts.extend(contracts)
        print(f"Total acumulado: {len(all_contracts)} contratos")
        
        # Salvar checkpoint por ano
        if contracts:
            year_df = pd.DataFrame(normalize_data(contracts))
            checkpoint_path = os.path.join("checkpoints", f"contratos_pncp_{year}.pkl")
            os.makedirs("checkpoints", exist_ok=True)
            year_df.to_pickle(checkpoint_path)
            print(f"Dados do ano {year} salvos em checkpoint.")
    
    if not all_contracts:
        print("Nenhum contrato encontrado")
        return
    
    # Normalizar dados (transformar estruturas aninhadas em colunas planas)
    print("Normalizando dados...")
    normalized_contracts = normalize_data(all_contracts)
    
    # Criar DataFrame
    df = pd.DataFrame(normalized_contracts)
    
    # Definir o caminho e o nome do arquivo
    PATH = "C:\\Users\\Haroldo Duraes\\Desktop\\GOvGO\\v0\\#DATA\\PNCP\\CONTRATOS\\"
    FILE = "contratos_pncp.xlsx"
    
    # Salvar em batches no arquivo Excel (1000 linhas por batch = 10 páginas)
    save_to_excel_in_batches(df, PATH, FILE, batch_size=1000)
    
    print(f"Processamento concluído. Total de {len(df)} contratos exportados.")

if __name__ == "__main__":
    main()