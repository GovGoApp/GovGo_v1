import requests
import pandas as pd
import os
import datetime
import concurrent.futures
import re
import json
import threading
from rich.progress import Progress, BarColumn, TextColumn, TimeElapsedColumn, TimeRemainingColumn
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Obter a data atual para nomear os arquivos
today = datetime.date.today()
data_str = today.strftime("%Y%m%d")
#data_str = '20250526'  # Para testes, use uma data fixa

# Caminhos e nomes de arquivos baseados na data atual
PATH = "C:\\Users\\Haroldo Duraes\\Desktop\\GOvGO\\v0\\#DATA\\PNCP\\"
IN_PATH = PATH + "DB\\CONTRATAÇÕES\\NEW\\"
OUT_PATH = PATH + "DB\\ITENS\\NEW\\"
IN_FILE = f"CONTRATAÇÕES_PNCP_{data_str}.xlsx"
OUT_FILE = f"ITENS_CONTRATAÇÕES_PNCP_{data_str}.xlsx"

# Variáveis globais para escrita imediata
workbook_lock = threading.Lock()
# Cria um novo workbook e remove a folha padrão
wb = Workbook()
wb.remove(wb.active)

# Lista de campos dos itens conforme o exemplo do arquivo itens.json (fixa)
ITEM_FIELDS = [
    "numeroItem",
    "descricao",
    "materialOuServico",
    "materialOuServicoNome",
    "valorUnitarioEstimado",
    "valorTotal",
    "quantidade",
    "unidadeMedida",
    "orcamentoSigiloso",
    "itemCategoriaId",
    "itemCategoriaNome",
    "patrimonio",
    "codigoRegistroImobiliario",
    "criterioJulgamentoId",
    "criterioJulgamentoNome",
    "situacaoCompraItem",
    "situacaoCompraItemNome",
    "tipoBeneficio",
    "tipoBeneficioNome",
    "incentivoProdutivoBasico",
    "dataInclusao",
    "dataAtualizacao",
    "temResultado",
    "imagem",
    "aplicabilidadeMargemPreferenciaNormal",
    "aplicabilidadeMargemPreferenciaAdicional",
    "percentualMargemPreferenciaNormal",
    "percentualMargemPreferenciaAdicional",
    "ncmNbsCodigo",
    "ncmNbsDescricao",
    "catalogo",
    "categoriaItemCatalogo",
    "catalogoCodigoItem",
    "informacaoComplementar"
]

def remove_illegal_chars(value):
    if isinstance(value, str):
        return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', value)
    return value

def clean_dataframe(df):
    return df.apply(lambda col: col.map(remove_illegal_chars))

# Processa uma linha usando apenas o campo "numeroControlePNCP"
def process_row(row):
    numeroControle = str(row.get("numeroControlePNCP", "")).strip()
    if not numeroControle:
        return []
    try:
        # Tentar reconhecer o formato pelo padrão esperado
        if not re.match(r'^\d+-\d+-\d+/\d+$', numeroControle):
            # Formato inválido, pular silenciosamente
            return []
            
        parts = numeroControle.split("-")
        if len(parts) != 3:
            return []
        cnpj = parts[0]
        # Ignora o prefixo (parts[1])
        seq_and_year = parts[2].split("/")
        if len(seq_and_year) != 2:
            return []
        seq = seq_and_year[0]
        anoCompra = seq_and_year[1]
        # Remove zeros à esquerda do sequencial
        sequencialCompra = str(int(seq))
        # Monta a URL:
        url = f"https://pncp.gov.br/api/pncp//v1/orgaos/{cnpj}/compras/{anoCompra}/{sequencialCompra}/itens"
        # Definindo timeout para evitar travamentos
        response = requests.get(url, timeout=20)
        if response.status_code != 200:
            return []
        itens = response.json()  # espera-se que seja uma lista
        resultados = []
        for item in itens:
            registro = {"numeroControlePNCP": numeroControle}
            for campo in ITEM_FIELDS:
                valor = item.get(campo)
                # Converter dicionários e listas para string JSON
                if isinstance(valor, (dict, list)):
                    registro[campo] = json.dumps(valor, ensure_ascii=False)
                else:
                    registro[campo] = valor
            resultados.append(registro)
        return resultados
    except requests.exceptions.Timeout:
        return []
    except Exception as e:
        return [{"numeroControlePNCP": numeroControle, "erro": str(e)}]
    
# Processa uma aba (sheet) do IN_FILE
def process_sheet(sheet_name, df, progress):
    # Remove linhas totalmente vazias
    df = df.dropna(how="all")
    progress.console.log(f"Aba '{sheet_name}' lida com {len(df)} linhas.")
    
    total_rows = len(df)
    sheet_task_id = progress.add_task(f"[bold blue]Aba {sheet_name}", total=total_rows)
    
    # Dividir o processamento em blocos de 1000 linhas
    BATCH_SIZE = 1000
    resultados_totais = 0
    
    for start_idx in range(0, total_rows, BATCH_SIZE):
        end_idx = min(start_idx + BATCH_SIZE, total_rows)
        batch_df = df.iloc[start_idx:end_idx]
        batch_size = len(batch_df)
        
        # Adicionar barra para o batch atual
        batch_id = progress.add_task(
            f"[bold yellow]  Batch {start_idx//BATCH_SIZE + 1} ({start_idx+1}-{end_idx})", 
            total=batch_size
        )
        
        resultados = []
        
        # Processar o lote atual
        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            future_to_index = {executor.submit(process_row, row): idx for idx, row in batch_df.iterrows()}
            for future in concurrent.futures.as_completed(future_to_index):
                try:
                    linhas = future.result()
                    for linha in linhas:
                        if 'erro' in linha:
                            progress.console.log(f"[red]Erro linha '{linha['numeroControlePNCP']}': {linha['erro']}")
                        else:
                            resultados.append(linha)
                except Exception as e:
                    progress.console.log(f"[red]Erro inesperado na execução da thread: {str(e)}")
                
                # Atualizar ambas as barras de progresso
                progress.update(batch_id, advance=1)
                progress.update(sheet_task_id, advance=1)
        
        # Criar DataFrame para o lote atual
        colunas_saida = ["numeroControlePNCP"] + ITEM_FIELDS
        if resultados:
            batch_result_df = pd.DataFrame(resultados)
            batch_result_df = clean_dataframe(batch_result_df)
            batch_result_df = batch_result_df.reindex(columns=colunas_saida)
            
            # Salvar o lote atual e liberar memória
            append_to_sheet(sheet_name, batch_result_df, OUT_PATH + OUT_FILE, progress)
            resultados_totais += len(batch_result_df)
            
            # Remover a barra do batch concluído
            progress.update(batch_id, completed=batch_size, 
                description=f"[bold green]  Batch {start_idx//BATCH_SIZE + 1:02d} ✓ {len(batch_result_df)} itens")
            
            # Liberar memória
            del batch_result_df
            del resultados
        else:
            progress.remove_task(batch_id)
            progress.console.log(f"[yellow]⚠ Batch {start_idx//BATCH_SIZE + 1}: Sem resultados")
    
    progress.remove_task(sheet_task_id)
    progress.console.log(f"[bold green]✅ Concluída a aba {sheet_name} com {resultados_totais} itens")
    
    # Retornar uma estrutura vazia apenas para manter compatibilidade
    return pd.DataFrame(columns=colunas_saida)

# Função para acrescentar dados em uma aba sem recarregar todo o arquivo
def append_to_sheet(sheet_name, df, output_file, progress):
    with workbook_lock:
        # Cria diretório de saída se não existir
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        
        # Verifica se o arquivo existe
        if not os.path.exists(output_file):
            # Se não existe, cria um novo arquivo
            temp_wb = Workbook()
            # Sempre remover a primeira aba em branco
            if "Sheet" in temp_wb.sheetnames:
                temp_wb.remove(temp_wb["Sheet"])
            ws = temp_wb.create_sheet(title=sheet_name)
            headers_added = False
        else:
            # Se existe, carrega o arquivo
            try:
                temp_wb = openpyxl.load_workbook(output_file)
                if sheet_name in temp_wb.sheetnames:
                    ws = temp_wb[sheet_name]
                    headers_added = True
                else:
                    ws = temp_wb.create_sheet(title=sheet_name)
                    headers_added = False
            except Exception as e:
                progress.console.log(f"[red]Erro ao carregar arquivo para append: {str(e)}")
                # Criar um novo arquivo se falhar ao carregar
                temp_wb = Workbook()
                if "Sheet" in temp_wb.sheetnames:
                    temp_wb.remove(temp_wb["Sheet"])
                ws = temp_wb.create_sheet(title=sheet_name)
                headers_added = False
        
        # Adiciona dados
        for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True)):
            if r_idx == 0 and headers_added:
                continue  # Pular cabeçalhos se já existirem
            ws.append(r)
            
        # Salva o arquivo
        try:
            temp_wb.save(output_file)
        except Exception as e:
            progress.console.log(f"[red]Erro ao salvar arquivo: {str(e)}")
            # Tenta salvar em um arquivo alternativo
            backup_file = output_file.replace('.xlsx', f'_backup_{datetime.datetime.now().strftime("%H%M%S")}.xlsx')
            progress.console.log(f"[yellow]Tentando salvar em arquivo alternativo: {backup_file}")
            temp_wb.save(backup_file)

def main():
    # Verifica se o arquivo de entrada existe
    input_file = IN_PATH + IN_FILE
    if not os.path.exists(input_file):
        print(f"[bold red]Erro: O arquivo de contratações do dia ({IN_FILE}) não foi encontrado em {IN_PATH}")
        print(f"Execute primeiro o script de download de contratações diárias.")
        return
    
    # Cria diretório de saída se não existir
    os.makedirs(OUT_PATH, exist_ok=True)
    
    # Lê o arquivo Excel com todas as abas
    xls = pd.ExcelFile(input_file)
    sheets = xls.sheet_names  # Lista de abas
    
    # Cria uma instância global de Progress (Rich) com colunas customizadas
    progress = Progress(
        TextColumn("[bold white]{task.description}"),
        BarColumn(complete_style="green", finished_style="bright_green"),
        "[progress.percentage]{task.percentage:>3.0f}%",
        "•",
        TimeElapsedColumn(),
        "•", 
        TimeRemainingColumn(),
    )
    
    with progress:
        # Task externa: Batch de abas
        outer_task = progress.add_task(f"[bold green]📊 Processando Itens de Contratações do dia {data_str}", total=len(sheets))
        
        progress.console.log(f"[bold cyan]🚀 Iniciando processamento de {len(sheets)} abas do arquivo diário")
        
        # Processar uma aba de cada vez (sequencialmente)
        for i, sheet in enumerate(sheets, 1):
            try:
                progress.console.log(f"[bold magenta]📋 Aba {i}/{len(sheets)}: {sheet}")
                df_sheet = pd.read_excel(xls, sheet_name=sheet)
                process_sheet(sheet, df_sheet, progress)
                # Liberar memória
                del df_sheet
            except Exception as e:
                progress.console.log(f"[bold red]❌ Erro ao processar a aba '{sheet}': {str(e)}")
            progress.update(outer_task, advance=1)
            
        progress.remove_task(outer_task)
        progress.console.log(f"[bold green]🏁 Processamento concluído!")
    
    print(f"\n📄 Planilha de itens do dia {data_str} salva em: {OUT_PATH + OUT_FILE}")

if __name__ == "__main__":
    main()