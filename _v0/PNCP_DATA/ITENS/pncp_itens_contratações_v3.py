import requests
import pandas as pd
import os
import datetime
import concurrent.futures
import re
import json
import threading
from rich.progress import Progress, BarColumn, TextColumn, TimeElapsedColumn, TimeRemainingColumn
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Lista de campos dos itens conforme o exemplo do arquivo itens.json (fixa)
ITEM_FIELDS = [
    "numeroItem",
    "descricao",
    "materialOuServico",
    "materialOuServicoNome",
    "valorUnitarioEstimado",
    "valorTotal",
    "quantidade",
    "unidadeMedida",
    "orcamentoSigiloso",
    "itemCategoriaId",
    "itemCategoriaNome",
    "patrimonio",
    "codigoRegistroImobiliario",
    "criterioJulgamentoId",
    "criterioJulgamentoNome",
    "situacaoCompraItem",
    "situacaoCompraItemNome",
    "tipoBeneficio",
    "tipoBeneficioNome",
    "incentivoProdutivoBasico",
    "dataInclusao",
    "dataAtualizacao",
    "temResultado",
    "imagem",
    "aplicabilidadeMargemPreferenciaNormal",
    "aplicabilidadeMargemPreferenciaAdicional",
    "percentualMargemPreferenciaNormal",
    "percentualMargemPreferenciaAdicional",
    "ncmNbsCodigo",
    "ncmNbsDescricao",
    "catalogo",
    "categoriaItemCatalogo",
    "catalogoCodigoItem",
    "informacaoComplementar"
]

def remove_illegal_chars(value):
    if isinstance(value, str):
        return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', value)
    return value

def clean_dataframe(df):
    return df.apply(lambda col: col.map(remove_illegal_chars))

# Processa uma linha usando apenas o campo "numeroControlePNCP".
def process_row(row):
    numeroControle = str(row.get("numeroControlePNCP", "")).strip()
    if not numeroControle:
        return []
    try:
        # Formato esperado: "00394452000103-1-000675/2021"
        parts = numeroControle.split("-")
        if len(parts) != 3:
            return []
        cnpj = parts[0]
        # Ignora o prefixo (parts[1])
        seq_and_year = parts[2].split("/")
        if len(seq_and_year) != 2:
            return []
        seq = seq_and_year[0]
        anoCompra = seq_and_year[1]
        # Remove zeros à esquerda do sequencial
        sequencialCompra = str(int(seq))
        # Monta a URL:
        url = f"https://pncp.gov.br/api/pncp//v1/orgaos/{cnpj}/compras/{anoCompra}/{sequencialCompra}/itens"
        # Definindo timeout para evitar travamentos
        response = requests.get(url, timeout=20)
        if response.status_code != 200:
            return []
        itens = response.json()  # espera-se que seja uma lista
        resultados = []
        for item in itens:
            registro = {"numeroControlePNCP": numeroControle}
            for campo in ITEM_FIELDS:
                registro[campo] = item.get(campo)
            resultados.append(registro)
        return resultados
    except requests.exceptions.Timeout:
        return []
    except Exception as e:
        return [{"numeroControlePNCP": numeroControle, "erro": str(e)}]

# Processa uma aba (sheet) do IN_FILE.
def process_sheet(sheet_name, df, progress):
    # Remove linhas totalmente vazias
    df = df.dropna(how="all")
    progress.console.log(f"Sheet '{sheet_name}' lida com {len(df)} linhas.")
    resultados = []
    total_rows = len(df)
    task_id = progress.add_task(f"[bold yellow]  {sheet_name} (contratações)", total=total_rows)

    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        future_to_index = {executor.submit(process_row, row): idx for idx, row in df.iterrows()}
        for future in concurrent.futures.as_completed(future_to_index):
            try:
                linhas = future.result()
                for linha in linhas:
                    if 'erro' in linha:
                        progress.console.log(f"[red]Erro linha '{linha['numeroControlePNCP']}': {linha['erro']}")
                    else:
                        resultados.append(linha)
            except Exception as e:
                progress.console.log(f"[red]Erro inesperado na execução da thread: {str(e)}")
            progress.update(task_id, advance=1)

    progress.remove_task(task_id)

    colunas_saida = ["numeroControlePNCP"] + ITEM_FIELDS
    if resultados:
        sheet_df = pd.DataFrame(resultados)
        sheet_df = clean_dataframe(sheet_df)
        sheet_df = sheet_df.reindex(columns=colunas_saida)
    else:
        sheet_df = pd.DataFrame(columns=colunas_saida)

    return sheet_df

# Variáveis globais para escrita imediata
workbook_lock = threading.Lock()
# Cria um novo workbook e remove a folha padrão.
wb = Workbook()
wb.remove(wb.active)

def write_sheet(sheet_name, df, output_file, progress):
    with workbook_lock:
        # Se a aba já existir, removê-la.
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            wb.remove(ws)
        ws = wb.create_sheet(title=sheet_name)
        from openpyxl.utils.dataframe import dataframe_to_rows
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        wb.save(output_file)
        progress.console.log(f"Sheet '{sheet_name}' gravada com {len(df)} linhas.")

def main():
    PATH = "C:\\Users\\Haroldo Duraes\\Desktop\\GOvGO\\v0\\#DATA\\PNCP\\"
    IN_PATH = PATH + "CONTRATAÇÕES\\"
    OUT_PATH = PATH + "ITENS\\"
    IN_FILE = "CONTRATAÇÕES_PNCP_03_2025.xlsx"
    OUT_FILE = "ITENS_CONTRATAÇÕES_PNCP_03_2025.xlsx"
    
    # Lê o arquivo Excel com todas as abas
    xls = pd.ExcelFile(IN_PATH + IN_FILE)
    sheets = xls.sheet_names  # Lista de abas
    
    # Cria uma instância global de Progress (Rich) com colunas customizadas
    progress = Progress(
        TextColumn("[bold yellow]{task.description}"),
        BarColumn(complete_style="green", finished_style="bright_green"),
        "[progress.percentage]{task.percentage:>3.0f}%",
        TimeElapsedColumn(),
        TimeRemainingColumn(),
    )
    
    with progress:
        # Task externa: Batch de abas
        outer_task = progress.add_task("[bold green]Batch mensal", total=len(sheets))
        with concurrent.futures.ThreadPoolExecutor(max_workers=13) as executor:
            future_to_sheet = {}
            for sheet in sheets:
                df_sheet = pd.read_excel(xls, sheet_name=sheet)
                future = executor.submit(process_sheet, sheet, df_sheet, progress)
                future_to_sheet[future] = sheet
            for future in concurrent.futures.as_completed(future_to_sheet):
                sheet = future_to_sheet[future]
                try:
                    sheet_df = future.result()
                    # Escreve imediatamente esta aba no arquivo de saída
                    write_sheet(sheet, sheet_df, OUT_PATH + OUT_FILE, progress)
                except Exception:
                    progress.console.log(f"[red]Erro ao processar a aba '{sheet}'.")
                progress.update(outer_task, advance=1)
        progress.remove_task(outer_task)
    
    print("\nPlanilha de itens consolidada salva em:", OUT_FILE)

if __name__ == "__main__":
    main()
